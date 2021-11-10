# this code is written by @yaAubakirov
# don't forget to install PyMuPDF, otherwise frontend error could appear
# PyMuPDF version should be not less than v1.19.1

import tkinter as tk
import tkinter.filedialog as dg
import tkinter.scrolledtext as st
import os
import sys
import re
import random
import threading
import openpyxl
import fitz


# this class is used as storage for variables to pass them between classes
class Storage:
    text = None
    weld_list = None
    ndt_list = None
    first_mark_list = None
    second_mark_list = None
    temp_drawing_number_list = None
    filename = None
    duplicated_welds = None
    list_of_found_welds = None

    @classmethod
    def clear_all(cls):
        Storage.text = None
        Storage.weld_list = None
        Storage.ndt_list = None
        Storage.first_mark_list = None
        Storage.second_mark_list = None
        Storage.temp_drawing_number_list = None
        Storage.filename = None
        Storage.duplicated_welds = None
        Storage.list_of_found_welds = None


# class which works with pdf
class Pdf:
    # this method extracts text from pdf
    @classmethod
    def extract_text_from_pdf(cls, pdf_path):
        all_text = ""
        with fitz.open(pdf_path) as doc:
            for num, page in enumerate(doc):
                all_text += page.get_text(clip=page.rect, sort=True)

        list_of_welds = Pdf.__find_all_welds(all_text)
        Storage.list_of_found_welds = list_of_welds
        Storage.text = all_text

    # finds all weld kind entities and pushes to list
    @classmethod
    def __find_all_welds(cls, text):
        # all welds to be extracted from plain text according to pattern below
        # there are many types of weld representation in drawing
        # w111A, 11111A, 11111 A etc.
        # to evade extraction special beam profiles as 128A or 136A
        # it is used dash along with w in the beginning of the pattern
        # and after such entities are deleted by list comprehension with condition
        pattern = r"\b[-w]?[^0T_-][\d]+[  ]?[A-D][\n]+\b"
        welds = re.findall(pattern, text)
        welds = list(set(welds))
        for i, weld in enumerate(welds):
            welds[i] = weld.replace("\n", "").replace(" ", "").replace(" ]", "").replace("\xa0", "")
        welds.sort()
        welds = [x for x in welds if x[0] != "-"]
        return welds


# class which analyzes the text
class Analyze:
    # this class gets weld number, its index and text. Looks for concatenated weld number with ndt class in text
    @classmethod
    def find_in_text(cls, to_find, index, text):
        # list with three types of concat
        ndt_classes = [
            ' ' + str(Storage.ndt_list[index]),
            str(Storage.ndt_list[index]),
            ' ' + str(Storage.ndt_list[index])
        ]
        for ndt in ndt_classes:
            try:
                res_search = re.search("".join([to_find, ndt]), text)
            except:
                return False
            if res_search:
                return True

    @classmethod
    def __weld_without_ndt(cls, weld, text):
        # if there are typical welds, this method is used to find just weld number
        try:
            res_search = re.search(weld, text)
        except:
            return False
        if res_search:
            return True

    @classmethod
    def is_weld_is_plating_grating(cls, weld, index, text):
        # method to understand if it is plating
        if Analyze.__weld_without_ndt(weld, text):
            if Storage.first_mark_list[index][:3] == "FLP" \
                    or Storage.second_mark_list[index][:3] == "FLP":
                return True
        else:
            return False

    @classmethod
    def is_weld_platform_plating(cls, weld, index, text):
        # method to understand if it is platform
        if Analyze.__weld_without_ndt(weld, text):
            if Storage.first_mark_list[index][:2] == "PL" and Storage.first_mark_list[index][7:9] == "PL":
                return True
            elif Storage.second_mark_list[index][:2] == "PL" and Storage.second_mark_list[index][7:9] == "PL":
                return True
            else:
                return False


class Excel:
    # this method extracts column and returns it as a list
    # it takes worksheet (ws), last_row (max_rows), column to extract as int (column), and row to start
    # if value is missed for cell it writes to list "missed value in row" with row number
    @classmethod
    def extract_from_sheet(cls, ws, max_rows, column, row=2):
        temp_list = []
        row_index = row
        for row in ws.iter_rows(min_row=row, min_col=column, max_col=column, max_row=max_rows):
            for cell in row:
                if cell.value:
                    if isinstance(cell.value, int):
                        temp_list.append(cell.value)
                    else:
                        stripped_value = cell.value.strip()
                        temp_list.append(stripped_value)
                else:
                    temp_list.append("missed value in row {}".format(row_index))
            row_index += 1
        return temp_list

    # this method is used to check if there are duplicated welds in WSL
    @classmethod
    def checking_welds_for_duplicates(cls, welds_list):
        seen = set()
        unique = [x for x in welds_list if x in seen and isinstance(x, int) or seen.add(x)]
        return unique


# main class which one runs application interface
class App:
    def __init__(self, master):
        version = 1.54
        # this is app icon file
        datafile = "my.ico"
        if not hasattr(sys, "frozen"):
            datafile = os.path.join(os.path.dirname(__file__), datafile)
        else:
            datafile = os.path.join(sys.prefix, datafile)

        # application interface
        self.master = master
        master.title('Welds checker (v{})'.format(version))
        master.geometry("600x400")
        master.resizable(0, 0)
        master.columnconfigure(2, weight=2)
        master.iconbitmap(default=datafile)

        self.load_pdf = tk.Button(master, text='PDF', height=1, width=10, bd=1, command=self.pdf_load)
        self.load_pdf.grid(row=1, column=2, pady=4, padx=4)

        self.load_excel = tk.Button(master, text='WSL', height=1, width=10, bd=1, command=self.excel_load)
        self.load_excel.grid(row=2, column=2, pady=4, padx=4)

        self.red_button = tk.Button(master, text='Check', height=1, width=10, bd=1, command=self.analyze)
        self.red_button.grid(row=3, column=2, pady=4, padx=4)

        self.txt = st.ScrolledText(master, width=40)
        self.txt.grid(rowspan=5, column=0, row=0, pady=4, padx=4)
        self.txt.configure(state='disabled')

        self.copyright = tk.Label(master, text='Metal Yapı Engineering & Construction LLC', fg="#808080")
        self.copyright.place(relx=.6, rely=.95)

    def pdf_load(self):
        self.clear_all_text()
        Storage.clear_all()
        file = dg.askopenfile(mode='rb', title='Choose a file', filetypes=[("PDF files", ".pdf")])
        if file is not None:
            filepath = os.path.abspath(file.name)
            filename = os.path.splitext(os.path.basename(filepath))[0]
            # to check if filename in appropriate format
            if "079322C-AWP1B-" not in filename[:14] and "079322C-GWP5B-" not in filename[:14]:
                self.insert_text("PDF file should be in appropriate format")
                self.insert_text("079322C-XXXXX-XXX-CS-KMD-XXXXX-XX-XXX")
                return False
            pdf_load = threading.Thread(target=Pdf.extract_text_from_pdf, args=[filepath])
            pdf_load.daemon = True
            pdf_load.start()
            while pdf_load.is_alive():
                self.loading()
            self.insert_text('Drawing is uploaded')
            Storage.filename = filename
        else:
            self.insert_text('PDF is not uploaded')
            return False

    def excel_load(self):
        try:
            file = dg.askopenfile(mode='rb', title='Choose WSL report', filetypes=[("Excel files", ".xlsx")])
        except PermissionError:
            self.insert_text('Close WSL first')
            return False
        if file is not None:
            filepath = os.path.abspath(file.name)
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            max_rows = ws.max_row
            Storage.weld_list = Excel.extract_from_sheet(ws, max_rows, 12)
            Storage.ndt_list = Excel.extract_from_sheet(ws, max_rows, 20)
            Storage.temp_drawing_number_list = Excel.extract_from_sheet(ws, max_rows, 4)
            Storage.first_mark_list = Excel.extract_from_sheet(ws, max_rows, 6)
            Storage.second_mark_list = Excel.extract_from_sheet(ws, max_rows, 9)
        else:
            self.insert_text('WSL is not uploaded')
            return False
        self.insert_text('WSL is uploaded')
        list_of_duplicates = list(set(Excel.checking_welds_for_duplicates(Storage.weld_list)))
        Storage.duplicated_welds = list_of_duplicates

    def analyze(self):
        wrong_welds = []
        typical_welds = []
        self.insert_text('.............')
        text = ""
        duplicated_welds = []
        if Storage.duplicated_welds:
            duplicated_welds = Storage.duplicated_welds
        if Storage.text is None:
            self.insert_text('PDF is not loaded')
        else:
            text = Storage.text

        if Storage.weld_list:
            for index, weld in enumerate(Storage.weld_list):
                if Analyze.find_in_text(str(weld), index, text):
                    if Storage.temp_drawing_number_list[index] == Storage.filename[:37]:
                        self.weld_text_insert(weld)
                    else:
                        wrong_welds.append(weld)
                        description = "drawing number for {}".format(weld)
                        self.problem_weld_text_insert(description)
                elif Analyze.is_weld_is_plating_grating(str(weld), index, text):
                    self.typical_weld_text_insert(weld)
                    typical_welds.append(weld)
                elif Analyze.is_weld_platform_plating(str(weld), index, text):
                    self.typical_weld_text_insert(weld)
                    typical_welds.append(weld)
                else:
                    self.problem_weld_text_insert(weld)
                    wrong_welds.append(weld)
        else:
            self.insert_text('WSL is not loaded')

        self.insert_text('.............')
        final_result = "\nTotal count of welds is {}".format(len(Storage.weld_list))
        self.insert_text(final_result)
        if len(typical_welds) > 0:
            final_result_for_typical_welds = "Total count of typical welds is {}".format(len(typical_welds))
            self.insert_text(final_result_for_typical_welds)
        if len(duplicated_welds) > 0:
            final_result_for_duplicated_welds = "Total count of duplicated welds is {}".format(len(duplicated_welds))
            self.insert_text(final_result_for_duplicated_welds)
        final_result_for_wrong_welds = "Total count of problem welds is {}".format(len(wrong_welds))
        self.insert_text(final_result_for_wrong_welds, 2)

        list_of_welds = Storage.weld_list
        found_welds = Storage.list_of_found_welds
        if len(found_welds) > 0 and (found_welds[0][0] == "w" or found_welds[len(found_welds) - 1][0] == "w"):
            found_welds = [weld.replace("w", "") for weld in found_welds]
        found_welds = [int(weld[:len(weld) - 1]) for weld in found_welds]
        spare_welds = list(set(found_welds).difference(set(list_of_welds)))

        if len(wrong_welds) == 0 and len(duplicated_welds) == 0 and len(spare_welds) == 0:
            self.insert_text('.............', 2)
            self.txt.configure(state='normal')
            self.txt.insert('end', "{} ✔\n".format(Storage.filename[:37]), 'name')
            self.txt.tag_config('name', foreground='green')
            self.txt.yview('end')
            self.txt.configure(state='disabled')
        elif len(duplicated_welds) > 0:
            self.insert_text('.............', 2)
            self.txt.configure(state='normal')
            self.txt.insert('end', "{} ✘\n".format(Storage.filename[:37]), 'warning')
            self.txt.tag_config('warning', foreground="red")
            self.txt.yview('end')
            self.txt.configure(state='disabled')
        elif len(wrong_welds) > 0:
            self.insert_text('.............', 2)
            self.txt.configure(state='normal')
            self.txt.insert('end', "{} ✘\n".format(Storage.filename[:37]), 'warning')
            self.txt.tag_config('warning', foreground="red")
            self.txt.yview('end')
            self.txt.configure(state='disabled')
        elif len(spare_welds) > 0 and len(wrong_welds) == 0:
            self.insert_text('.............', 2)
            self.txt.configure(state='normal')
            self.txt.insert('end', "{} ✔\n".format(Storage.filename[:37]), 'org')
            self.txt.tag_config('org', foreground='orange')
            self.txt.yview('end')
            self.txt.configure(state='disabled')
        else:
            self.insert_text('.............', 2)
            self.txt.configure(state='normal')
            self.txt.insert('end', "{} ✔\n".format(Storage.filename[:37]), 'org')
            self.txt.tag_config('org', foreground='orange')
            self.txt.yview('end')
            self.txt.configure(state='disabled')
        if len(Storage.weld_list) == len(wrong_welds):
            self.insert_text('\n.............')
            self.insert_text('Probably spaces are not deleted from WSL')
            self.insert_text('Or you have chosen wrong WSL')
            return 0
        self.txt.yview('end')

        if len(wrong_welds) > 0:
            self.insert_text('\n.............')
            self.insert_text('Problem welds:')
            for weld in wrong_welds:
                self.insert_text(weld)

        if len(Storage.duplicated_welds) > 0:
            self.insert_text('\n.............')
            self.insert_text('Duplicated welds:')
            for weld in Storage.duplicated_welds:
                self.insert_text(weld)

        if len(spare_welds) > 0:
            self.insert_text('\n.............')
            self.insert_text('These welds are not presented in WSL:')
            for weld in spare_welds:
                self.insert_text(weld)

    def loading(self):
        hashtags = random.randint(5, 35)
        spaces = 40 - hashtags - 4
        percentage = int(hashtags * 2.5)
        self.txt.configure(state='normal')
        self.txt.insert('end', 'Pdf pages loading. Please wait...\n')
        self.txt.insert('end', '#' * hashtags + ' ' * spaces + '{}%'.format(percentage))
        self.refresh()
        self.clear_all_text()
        self.txt.configure(state='disabled')

    def insert_text(self, text, number_of_n=1):
        self.txt.configure(state='normal')
        self.txt.insert('end', "{}{}".format(text, '\n' * number_of_n))
        self.txt.yview('end')
        self.txt.configure(state='disabled')

    def weld_text_insert(self, weld_number):
        self.txt.configure(state='normal')
        self.refresh()
        self.txt.insert('end', "{} is OK\n".format(str(weld_number)))
        self.txt.yview('end')
        self.txt.configure(state='disabled')

    def problem_weld_text_insert(self, weld_number):
        self.txt.configure(state='normal')
        self.refresh()
        self.txt.insert('end', "Problem with {}\n".format(str(weld_number)), 'warning')
        self.txt.tag_config('warning', foreground="red")
        self.txt.yview('end')
        self.txt.configure(state='disabled')

    def typical_weld_text_insert(self, weld_number):
        self.txt.configure(state='normal')
        self.refresh()
        self.txt.insert('end', "{} is typical\n".format(str(weld_number)), 'typical')
        self.txt.tag_config('typical', foreground='orange')
        self.txt.yview('end')
        self.txt.configure(state='disabled')

    def refresh(self):
        self.master.update()

    def clear_all_text(self):
        self.txt.delete('1.0', tk.END)


root = tk.Tk()
my_gui = App(root)
root.mainloop()
